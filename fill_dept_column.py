"""
Excel表格部门列填充工具

功能：
1. 新建第1列（部门）、第2列（职务）
2. 识别部门标题行（第1列有部门名称，第2列为空）
3. 将部门名称填充到其下属人员行的新第1列
4. 删除姓名列中的空格
5. 删除部门标题行
6. 保存为新文件

使用方法：
    skill: fill_dept_column 学院通讯录.xlsx
    skill: fill_dept_column input.xlsx output.xlsx
"""

import openpyxl
import sys
import os

# 设置UTF-8编码输出
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')


def fill_dept_column(input_file, output_file=None):
    """
    处理Excel文件，新建部门列并删除部门标题行

    Args:
        input_file: 输入Excel文件路径
        output_file: 输出Excel文件路径（可选，默认为原文件名_v2.xlsx）
    """
    if not os.path.exists(input_file):
        print(f"错误：文件不存在 - {input_file}")
        return False

    if output_file is None:
        base_name = os.path.splitext(input_file)[0]
        output_file = f"{base_name}_v2.xlsx"

    print(f"读取文件：{input_file}")
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    # 创建新工作簿
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active

    # 第1行：保持原标题，前面加空列
    new_ws.cell(row=1, column=1).value = None
    for col_idx, cell in enumerate(ws[1], 2):
        new_ws.cell(row=1, column=col_idx).value = cell.value

    # 第2行：设置新标题（跳过原表第2行的第1列"部门或职务"）
    new_ws.cell(row=2, column=1).value = '部门'
    new_ws.cell(row=2, column=2).value = '职务'
    for col_idx, cell in enumerate(ws[2][1:], 3):
        new_ws.cell(row=2, column=col_idx).value = cell.value

    # 处理数据行（从原表第3行开始）
    current_dept = None
    new_row_idx = 3
    deleted_rows = 0

    for row_idx in range(3, ws.max_row + 1):
        # 读取原行
        orig_row = [cell.value for cell in ws[row_idx]]

        # 判断是否是部门标题行：第2列（姓名）为None
        if orig_row[1] is None:
            # 这是部门标题行，记录部门名称，跳过
            current_dept = orig_row[0]
            deleted_rows += 1
            continue
        else:
            # 这是人员行，构建新行：第1列=部门，第2列=职务，后面是原数据
            # 删除姓名列（第3列，即原表的第2列）中的空格
            name_value = orig_row[1]
            if name_value and isinstance(name_value, str):
                name_value = name_value.replace(' ', '')
            new_row = [current_dept, orig_row[0]] + [name_value] + orig_row[2:]
            for col_idx, value in enumerate(new_row, 1):
                new_ws.cell(row=new_row_idx, column=col_idx).value = value
            new_row_idx += 1

    # 保存新文件
    new_wb.save(output_file)
    print(f"[OK] 处理完成！")
    print(f"  原表行数：{ws.max_row}")
    print(f"  新表行数：{new_row_idx - 1}")
    print(f"  删除行数：{deleted_rows}")
    print(f"  保存到：{output_file}")

    return True


def main(args):
    """
    Skill入口函数

    Args:
        args: 命令行参数
            args[0]: 输入Excel文件路径
            args[1]: 输出Excel文件路径（可选）
    """
    if len(args) < 1:
        print("使用方法：")
        print("  skill: fill_dept_column <输入文件>")
        print("  skill: fill_dept_column <输入文件> <输出文件>")
        print()
        print("示例：")
        print("  skill: fill_dept_column 学院通讯录.xlsx")
        print("  skill: fill_dept_column data.xlsx result.xlsx")
        return

    input_file = args[0]
    output_file = args[1] if len(args) > 1 else None

    fill_dept_column(input_file, output_file)


if __name__ == "__main__":
    main(sys.argv[1:])
